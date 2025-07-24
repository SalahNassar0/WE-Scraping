import asyncio
import os
import re
import sys
import logging
from dotenv import load_dotenv
from playwright.async_api import async_playwright
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from datetime import datetime, timedelta
import yagmail
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
import shutil
from pathlib import Path

def archive_daily_report():
    """Archive today's report with date in filename"""
    from datetime import datetime
    import shutil
    
    today_str = datetime.now().strftime("%Y-%m-%d")
    archive_path = f"usage_report_{today_str}.xlsx"
    
    try:
        shutil.copyfile("usage_report.xlsx", archive_path)
        logging.info(f"Archived daily report as {archive_path}")
    except Exception as e:
        logging.error(f"Failed to archive daily report: {e}")

def save_to_csv(df_for_excel):
    """Save daily report to CSV for dashboard"""
    csv_path = Path("daily_reports.csv")
    
    # Add date column
    df_for_excel['date'] = pd.Timestamp.now().date()
    
    # Append to CSV (create if doesn't exist)
    if csv_path.exists():
        df_for_excel.to_csv(csv_path, mode='a', header=False, index=False)
    else:
        df_for_excel.to_csv(csv_path, mode='w', header=True, index=False)
    
    logging.info(f"Daily report saved to {csv_path}")

# ── Configure Logging ────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(message)s',
    handlers=[
        logging.FileHandler("scraper.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

load_dotenv()

# ── Browser Launch Arguments ─────────────────────────────────────────────────
BROWSER_LAUNCH_ARGS = [
    '--no-sandbox',
    '--disable-setuid-sandbox',
    '--disable-dev-shm-usage',
    '--disable-gpu',
]

# ── Slack settings ───────────────────────────────────────────────────────────
SLACK_BOT_TOKEN = os.getenv("SLACK_BOT_TOKEN")
SLACK_CHANNEL_ID = os.getenv("SLACK_CHANNEL_ID")

if not (SLACK_BOT_TOKEN and SLACK_CHANNEL_ID):
    logging.warning("Slack Bot Token or Channel ID not found in .env. Slack alerts will be disabled.")

# ── Email settings ────────────────────────────────────────────────────────────
EMAIL_SENDER     = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD   = os.getenv("EMAIL_PASSWORD")
EMAIL_RECIPIENT  = os.getenv("EMAIL_RECIPIENT", "").strip()
EMAIL_RECIPIENTS = [
    addr.strip()
    for addr in os.getenv("EMAIL_RECIPIENTS", "").split(",")
    if addr.strip()
]
TO_ADDRS = EMAIL_RECIPIENTS or ([EMAIL_RECIPIENT] if EMAIL_RECIPIENT else [])


# ── Conditional Formatting Thresholds ────────────────────────────────────────
LOW_REMAINING_YELLOW_GB = float(os.getenv("LOW_REMAINING_YELLOW_GB", "80"))
LOW_REMAINING_RED_GB    = float(os.getenv("LOW_REMAINING_RED_GB", "20"))


# ── Build accounts from .env ─────────────────────────────────────────────────
accounts = []
i = 1
while os.getenv(f"ACCOUNT{i}_PHONE"):
    accounts.append({
        "phone":    os.getenv(f"ACCOUNT{i}_PHONE"),
        "password": os.getenv(f"ACCOUNT{i}_PASS"),
        "type":     os.getenv(f"ACCOUNT{i}_TYPE", "Internet"),
        "store":    os.getenv(f"ACCOUNT{i}_NAME", f"Account {i}")
    })
    i += 1
if not accounts:
    logging.warning("No accounts configured! Please set ACCOUNT1_PHONE, ACCOUNT1_PASS, etc., in .env")
    

# ── Scraper ──────────────────────────────────
async def fetch_usage(account, browser):
    page = None
    context = None
    # Default payload, updated upon successful scraping
    result_payload = {
        "Store": account["store"], 
        "Number": account["phone"], 
        "Balance": "0 EGP (Not Found)",
        "Remaining": 0.0, "Used": 0.0, 
        "Add-ons": "N/A (No Details)", 
        "Add-ons Price": "N/A (No Details)",
        "Renewal Cost": "0 EGP (No Details)", 
        "Renewal Date": "(No Details)"
    }
    # Error payload structure if a critical, unrecoverable failure occurs for this account
    critical_error_payload = {
        "Store": account["store"], 
        "Number": account["phone"], 
        "Balance": "Error EGP", 
        "Remaining": 0.0, 
        "Used": 0.0, 
        "Add-ons": "Error",
        "Add-ons Price": "Error", 
        "Renewal Cost": "Error EGP",
        "Renewal Date": "Error"
    }

    try:
        logging.info(f"Processing: {account['store']} ({account['phone']})")
        context = await browser.new_context(
            user_agent=( "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        )
        page = await context.new_page()
        
        await page.goto("https://my.te.eg/echannel/#/login", timeout=60000)
        await page.fill('input[placeholder="Service number"]', account["phone"])
        await page.click(".ant-select-selector")
        await page.click(f'.ant-select-item-option >> text={account["type"]}')
        await page.fill('input[placeholder="Password"]', account["password"])
        await page.click('button:has-text("Login")')
        logging.info(f"Login initiated for {account['phone']}")

        await page.wait_for_load_state("networkidle", timeout=70000) 
        await page.wait_for_timeout(7000) 
        
        try:
            # Wait for the main dashboard content area to load
            await page.wait_for_selector('//span[normalize-space(text())="Current Balance"]/parent::div', timeout=30000) 
            logging.info(f"Main dashboard content area detected for {account['phone']}")
        except Exception as e:
            logging.error(f"Dashboard did not fully load or key element missing for {account['phone']}: {e}")
            raise TimeoutError(f"Dashboard key element timeout for {account['phone']}")

        balance_val = "0"; remaining_val = 0.0; used_val = 0.0
        
        # Balance
        try:
            bal_loc = page.locator('//span[normalize-space(text())="Current Balance"]/parent::div//div[contains(@style,"font-size")]').first
            await bal_loc.wait_for(timeout=15000, state="visible")
            txt = (await bal_loc.text_content() or "").strip().split()[0]
            if txt and txt != "0": balance_val = txt
            result_payload["Balance"] = f"{balance_val} EGP"; logging.info(f"Balance: {balance_val} for {account['phone']}")
        except Exception as e: logging.warning(f"Could not scrape Balance for {account['phone']}: {e}")

        # Remaining
        try:
            rem_loc = page.locator('//span[contains(.,"Remaining")]/preceding-sibling::span[1]').first
            await rem_loc.wait_for(timeout=15000, state="visible")
            val_str = (await rem_loc.text_content() or "0").strip()
            remaining_val = float(re.sub(r'[^\d.]', '', val_str) or 0.0)
            result_payload["Remaining"] = remaining_val; logging.info(f"Remaining: {remaining_val} for {account['phone']}")
        except Exception as e: logging.warning(f"Could not scrape Remaining for {account['phone']}: {e}")

        # Used
        try:
            used_loc = page.locator('//span[contains(.,"Used")]/preceding-sibling::span[1]').first
            await used_loc.wait_for(timeout=15000, state="visible")
            val_str = (await used_loc.text_content() or "0").strip()
            used_val = float(re.sub(r'[^\d.]', '', val_str) or 0.0)
            result_payload["Used"] = used_val; logging.info(f"Used: {used_val} for {account['phone']}")
        except Exception as e:
            logging.warning(f"Could not scrape Used for {account['phone']}: {e}")
            if not (await page.locator('//span[contains(.,"Used")]/preceding-sibling::span[1]').first.is_visible(timeout=1000)):
                logging.info(f"'Used' element not found for {account['phone']}")
        
        renewal_cost_val = "0"; renewal_date_val = ""; addon_names_val = "N/A"; addon_prices_val = "N/A"
        try: 
            more_details_locators = [page.locator('//span[text()="More Details"]').first, page.locator('//a[.//span[contains(text(),"details")]] | //button[.//span[contains(text(),"details")]]').first]
            more_details_clicked = False
            for md_loc in more_details_locators:
                try: await md_loc.wait_for(state='visible', timeout=10000); await md_loc.click(timeout=5000); more_details_clicked = True; logging.info(f"Clicked 'More Details' for {account['phone']}"); break
                except Exception: logging.debug(f"More details locator variant failed for {account['phone']}.")
            
            if not more_details_clicked: logging.warning(f"Could not click 'More Details' for {account['phone']}.")
            else:
                await page.wait_for_load_state("networkidle", timeout=45000); await page.wait_for_timeout(5000)
                addon_names_list = []; addon_prices_list_scraped = []
                not_subscribed_locator = page.locator('//span[contains(normalize-space(.), "You are not subscribed to any bundles currently")]')
                try: await not_subscribed_locator.wait_for(timeout=3000, state="visible"); logging.info(f"Not subscribed to bundles for {account['phone']}.")
                except Exception:
                    logging.info(f"Checking for add-on cards for {account['phone']}.")
                    addon_card_selector = ('//div[contains(@class, "slick-slide") and @aria-hidden="false"]//div[contains(@style, "border-style: solid")]')
                    addon_cards = page.locator(addon_card_selector); num_addon_cards = await addon_cards.count()
                    if num_addon_cards > 0: logging.info(f"Found {num_addon_cards} add-on cards for {account['phone']}.")
                    for i in range(num_addon_cards):
                        card = addon_cards.nth(i); name_text_cleaned = "N/A"; price_text_raw = "N/A"
                        try: name_loc = card.locator('xpath=.//div[contains(@style, "font-weight: bold;")]').first; await name_loc.wait_for(timeout=3000); name_text_original = (await name_loc.text_content() or "").strip(); name_text_cleaned = name_text_original if name_text_original else "N/A"
                        except Exception as e_name: logging.debug(f"Add-on name err for {account['phone']}: {e_name}")
                        try:
                            price_locator_xpath = './/span[contains(normalize-space(.), "Price:")]'; price_elements = card.locator(f"xpath={price_locator_xpath}"); 
                            if await price_elements.count() > 0:
                                actual_price_text = (await price_elements.first.text_content(timeout=3000) or "").strip()
                                if "Price:" in actual_price_text: parsed_price_val = actual_price_text.split("Price:", 1)[-1].strip(); price_text_raw = parsed_price_val if parsed_price_val else "N/A"
                                else: price_text_raw = "N/A"
                            else: price_text_raw = "N/A"
                        except Exception as e_price: logging.debug(f"Add-on price err for {account['phone']}: {e_price}")
                        if name_text_cleaned != "N/A" and price_text_raw != "N/A" and "EGP" in price_text_raw:
                            price_numeric_part = re.sub(r'[^\d.]', '', price_text_raw.split("EGP")[0])
                            if price_numeric_part and price_numeric_part in name_text_cleaned:
                                patterns = [r'-\s*' + re.escape(price_numeric_part) + r'\s*EGP\s*/\s*month',r'-\s*' + re.escape(price_numeric_part) + r'\s*EGP']
                                temp_name = name_text_cleaned; 
                                for pat in patterns: temp_name = re.sub(pat, '', temp_name, flags=re.IGNORECASE).strip()
                                name_text_cleaned = temp_name.rstrip('- /').strip()
                        if name_text_cleaned != "N/A" or price_text_raw != "N/A":
                            addon_names_list.append(name_text_cleaned if name_text_cleaned != "N/A" else "Unknown"); addon_prices_list_scraped.append(price_text_raw if price_text_raw != "N/A" else "0 EGP")
                    if addon_names_list: addon_names_val = "; ".join(addon_names_list); addon_prices_val = "; ".join(addon_prices_list_scraped)
                result_payload["Add-ons"] = addon_names_val; result_payload["Add-ons Price"] = addon_prices_val;
                logging.info(f"Add-ons: {addon_names_val}, Prices: {addon_prices_val} for {account['phone']}")

                try:
                    cost_loc = page.locator('//span[contains(text(),"Renewal Cost")]/following-sibling::span//div[1]').first
                    await cost_loc.wait_for(timeout=10000); val = (await cost_loc.text_content() or "").strip()
                    if val and val != "0": renewal_cost_val = val
                    logging.info(f"Renewal Cost: {renewal_cost_val} for {account['phone']}")
                except Exception as e: logging.warning(f"Could not scrape Renewal Cost for {account['phone']}: {e}")
                result_payload["Renewal Cost"] = f"{renewal_cost_val} EGP"
                
                try:
                    date_loc = page.locator('//span[contains(text(),"Renewal Date:")]').first
                    await date_loc.wait_for(timeout=10000); full_text_content = await date_loc.text_content() or ""
                    match = re.search(r"Renewal Date:\s*([\d-]+)", full_text_content)
                    if match: renewal_date_val = match.group(1).strip()
                    else: parts = full_text_content.split("Renewal Date:", 1); renewal_date_val = parts[1].split(",")[0].strip() if len(parts) > 1 else ""
                    logging.info(f"Renewal Date: {renewal_date_val} for {account['phone']}")
                except Exception as e: logging.warning(f"Could not parse Renewal Date for {account['phone']}: {e}")
                result_payload["Renewal Date"] = renewal_date_val
        except Exception as e: 
            logging.warning(f"Error in 'More Details' section or beyond for {account['phone']}: {e}")
        
        return result_payload
    except Exception as e: 
        logging.error(f"CRITICAL failure during fetch_usage for {account['store']} ({account['phone']}): {e}", exc_info=True)
        return critical_error_payload 
    finally:
        if page: 
            try: 
                await page.close()
            except Exception: pass
        if context: 
            try: 
                await context.close()
            except Exception: pass
        logging.info(f"Finished task for {account['phone']}.")

def parse_egp_string(cost_str):
    if not isinstance(cost_str, str) or "error" in cost_str.lower() or cost_str.lower() == "n/a" or "(not found)" in cost_str.lower() or "(no details)" in cost_str.lower():
        return 0.0
    total_value = 0.0; parts = cost_str.split(';')
    for part in parts:
        numeric_part = re.sub(r"[^\d\.]", "", part.strip())
        if numeric_part:
            try: total_value += float(numeric_part)
            except ValueError: logging.warning(f"Could not parse '{numeric_part}' from '{part}' to float.")
    return total_value

def send_slack_message(message_text):
    if not SLACK_BOT_TOKEN or not SLACK_CHANNEL_ID: logging.warning("Slack creds missing. No Slack msg."); return False
    try:
        client = WebClient(token=SLACK_BOT_TOKEN)
        client.chat_postMessage(channel=SLACK_CHANNEL_ID, text=message_text)
        logging.info(f"Slack message sent to {SLACK_CHANNEL_ID}."); return True
    except SlackApiError as e: logging.error(f"Error sending Slack message: {e.response['error']}"); return False

async def main():
    logging.info("Starting web scraping process...")
    if not accounts: 
        logging.warning("No accounts configured. Exiting.")
        return

    rows = []
    # AWS instance have low specs so we only use a single browser instance and limit concurrency
    async with async_playwright() as p:
        browser = None
        logging.info("Initializing Playwright...")
        try:
            logging.info("Playwright started. Launching single browser instance...")
            browser = await p.chromium.launch(headless=True, args=BROWSER_LAUNCH_ARGS)
            logging.info("Single browser instance launched.")
            
            CONCURRENCY_LIMIT = 3 # Adjust as needed (e.g., 1, 2, or the max your instance can handle)
            semaphore = asyncio.Semaphore(CONCURRENCY_LIMIT)

            async def limited_fetch_usage(account, browser, semaphore):
                async with semaphore:
                    logging.info(f"Semaphore acquired for {account['store']}, starting fetch_usage.")
                    result = await fetch_usage(account, browser)
                    logging.info(f"Semaphore released for {account['store']}.")
                    return result
            
            tasks = [limited_fetch_usage(ac, browser, semaphore) for ac in accounts]
            rows = await asyncio.gather(*tasks)
            
        except Exception as e: # Catches errors during browser setup or asyncio.gather itself
            logging.error(f"CRITICAL ERROR during Playwright browser setup or asyncio.gather: {e}", exc_info=True)
            critical_msg = (f":skull_and_crossbones: *SCRIPT EXECUTION FAILED!* ({pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')})\n"
                            "A critical error occurred with browser setup or task gathering. Script will exit.\nDetails: " + str(e))
            if SLACK_BOT_TOKEN and SLACK_CHANNEL_ID: send_slack_message(critical_msg)
            if TO_ADDRS and EMAIL_SENDER and EMAIL_PASSWORD:
                try:
                    yag = yagmail.SMTP(EMAIL_SENDER, EMAIL_PASSWORD)
                    yag.send(to=TO_ADDRS, subject="🚨 CRITICAL: WE Scraper Script Execution Failed!", contents=critical_msg)
                    if yag: yag.close()
                except Exception as email_err: logging.error(f"Failed to send critical failure email: {email_err}")
            return 
        finally:
            if browser and browser.is_connected():
                try: 
                    await browser.close()
                    logging.info("Browser instance closed.")
                except Exception as e_br_close: 
                    logging.error(f"Error closing browser: {e_br_close}")
            logging.info("Playwright operations finished.")
        
    logging.info(f"Finished scraping phase. Processed {len(rows)} accounts.")

    if not rows and accounts: #a safeguard
        logging.error("No data rows returned from scraping tasks, despite having accounts. Exiting.")
        if SLACK_BOT_TOKEN and SLACK_CHANNEL_ID:
            send_slack_message(":x: SCRIPT ERROR: No data rows returned from scraping tasks (gather might have failed or been cancelled early). Please check logs.")
        return
    elif not accounts: # Should have been caught at the very top of main()
        logging.info("No accounts were configured to process.")
        return 

    # --- Check if ALL accounts critically failed within fetch_usage ---
    critically_failed_accounts_count = 0
    if rows: # Ensure rows list is not None and has content
        for row_data in rows:
            # Define "critically failed" if Balance AND Add-ons indicate "Error" from fetch_usage's main except block
            if row_data.get("Balance") == "Error EGP" and row_data.get("Add-ons") == "Error":
                critically_failed_accounts_count += 1
    
    if len(rows) > 0 and critically_failed_accounts_count == len(rows):
        all_failed_msg = (
            f":rotating_light: *CRITICAL: All WE Accounts Failed Data Retrieval!* ({pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')})\n\n"
            f"The script could not fetch data for any of the {len(rows)} accounts.\n"
            f"This might be due to WE server issues, website changes, or network problems.\n\n"
            f"Please check the WE website status and review server logs (`scraper.log`) immediately. No report was generated."
        )
        logging.error(all_failed_msg)
        if SLACK_BOT_TOKEN and SLACK_CHANNEL_ID: send_slack_message(all_failed_msg)
        
        # Also email this critical "all accounts failed" alert if configured
        if TO_ADDRS and EMAIL_SENDER and EMAIL_PASSWORD:
            try:
                yag_fail = yagmail.SMTP(EMAIL_SENDER, EMAIL_PASSWORD)
                yag_fail.send(to=TO_ADDRS, subject="⚠️ URGENT: WE Scraper - All Accounts Failed Data Retrieval!", contents=all_failed_msg)
                if yag_fail: yag_fail.close()
            except Exception as email_err: 
                logging.error(f"Failed to send 'all accounts failed' email: {email_err}")
        
        logging.info("Exiting script due to critical failure on all accounts. No Excel report will be generated.")
        return # Stop further processing
    
    df = pd.DataFrame(rows)
    if df.empty and accounts: # Should be caught by 'rows' check, but as a safeguard
        logging.warning("DataFrame is empty after attempting to create it from rows. No report generated.")
        if SLACK_BOT_TOKEN and SLACK_CHANNEL_ID:
             send_slack_message(":warning: Script Warning: DataFrame was empty after processing. No report generated.")
        return

    # --- Process DataFrame for Calculations and Alerts (Numeric Columns) ---
    df["Balance Numeric"] = df["Balance"].apply(parse_egp_string)
    df["Renewal Cost Numeric"] = df["Renewal Cost"].apply(parse_egp_string)
    df["Add-ons Price Numeric"] = df["Add-ons Price"].apply(parse_egp_string)
    df["Total Cost Numeric"] = df["Renewal Cost Numeric"] + df["Add-ons Price Numeric"]
    df["Renewal Date DT"] = pd.to_datetime(df["Renewal Date"], format='%d-%m-%Y', errors='coerce')
    df["Remaining"] = pd.to_numeric(df["Remaining"], errors='coerce').fillna(0.0)
    
    if 'Used' in df.columns: 
        df["Used"] = pd.to_numeric(df["Used"], errors='coerce').fillna(0.0)
    else: # Should be present if fetch_usage always returns it
        df["Used"] = 0.0 
        logging.warning("'Used' column was missing from some DataFrame rows, defaulting to 0 for Main Quota calculation.")
    df["Main Quota"] = df["Remaining"] + df["Used"]
    logging.info("DataFrame numeric columns processed for alert logic.")

    # --- SLACK ALERTS AND REPORTING LOGIC ---
    individual_alerts_to_send = []
    low_gb_alert_count = 0
    renewal_low_balance_alert_count = 0
    
    # Mask for rows that were NOT critically failed (i.e., Balance is not "Error EGP")
    # Alerts should only be generated from these non-critically-failed rows.
    successful_scrape_mask = df["Balance"] != "Error EGP" 

    if SLACK_BOT_TOKEN and SLACK_CHANNEL_ID:
        logging.info("Gathering data for Slack alerts from non-critically-failed accounts..."); 
        for index, row in df[successful_scrape_mask].iterrows(): 
            if pd.notna(row['Remaining']) and row['Remaining'] < LOW_REMAINING_RED_GB :
                message = (f":warning: *Low GB Alert!* Acct: *{row['Store']}* ({row['Number']}) - Rem: *{row['Remaining']:.2f}GB*"); 
                individual_alerts_to_send.append(message); low_gb_alert_count += 1
        
        current_date_for_alerts = pd.Timestamp.now().normalize()
        for index, row in df[successful_scrape_mask].iterrows():
            if pd.notna(row['Renewal Date DT']) and \
               pd.notna(row['Balance Numeric']) and \
               pd.notna(row['Total Cost Numeric']) and \
               row['Renewal Cost'] != "Error EGP": # Additional check that renewal cost was actually scraped
                days_to_renewal = (row['Renewal Date DT'] - current_date_for_alerts).days
                if days_to_renewal <= 20 and row['Balance Numeric'] < row['Total Cost Numeric']:
                    message = (f":alarm_clock: *Low Balance!* Acct: *{row['Store']}* ({row['Number']}) - Renews: *{row['Renewal Date']}* ({days_to_renewal}d) - Balance: *{row['Balance Numeric']:.2f}*, Cost: *{row['Total Cost Numeric']:.2f}*"); 
                    individual_alerts_to_send.append(message); renewal_low_balance_alert_count += 1
        
        now = datetime.now()
        target_summary_time = now.replace(hour=12, minute=0, second=0, microsecond=0)
        interval_minutes = 10
        summary_window_start = target_summary_time - timedelta(minutes=interval_minutes)
        summary_window_end = target_summary_time + timedelta(minutes=interval_minutes)
        # now = datetime.now().replace(hour=12, minute=5) # Test summary window
        is_summary_time_window = summary_window_start <= now <= summary_window_end

        if is_summary_time_window:
            logging.info(f"Time {now.strftime('%H:%M')} is in 12PM summary window. Sending summary ONLY to Slack.")
            summary_parts = [f"*📊 Daily Usage Report Summary ({pd.Timestamp.now().strftime('%Y-%m-%d %I:%M %p')})*"]
            summary_parts.append(f"Total accounts processed: *{len(df)}* (Alerts based on *{successful_scrape_mask.sum()}* accounts with valid data)")
            if low_gb_alert_count > 0: summary_parts.append(f":warning: *{low_gb_alert_count} account(s)* have Low GB.")
            else: summary_parts.append(f":white_check_mark: No accounts with Low GB (among those successfully checked).")
            if renewal_low_balance_alert_count > 0: summary_parts.append(f":alarm_clock: *{renewal_low_balance_alert_count} account(s)* require Renewal/Low Balance attention.")
            else: summary_parts.append(f":white_check_mark: No Renewal/Low Balance concerns (among those successfully checked).")
            
            accounts_with_errors = len(df) - successful_scrape_mask.sum() # Count of "Error EGP" rows
            if accounts_with_errors > 0 :
                 summary_parts.append(f":exclamation: *{accounts_with_errors} account(s)* had critical data retrieval issues (marked 'Error' in report).")

            if successful_scrape_mask.sum() == len(df) and low_gb_alert_count == 0 and renewal_low_balance_alert_count == 0 : 
                summary_parts.append(f"\nOverall: :thumbsup: All accounts successfully processed and looking good!")
            
            if TO_ADDRS and EMAIL_SENDER and EMAIL_PASSWORD: 
                summary_parts.append(f"📧 _The detailed Excel report, containing full data for all accounts, has also been sent via email._")
            else: 
                summary_parts.append(f"📧 _Email reporting is not configured._")
            send_slack_message("\n".join(summary_parts))
            logging.info("Daily summary sent to Slack.")
        else: # Not summary window
            logging.info(f"Time {now.strftime('%H:%M')} is outside 12PM summary window. Sending individual/all-clear/mixed status.")
            if individual_alerts_to_send: 
                unique_alerts = sorted(list(set(individual_alerts_to_send)))
                logging.info(f"Sending {len(unique_alerts)} unique individual alert(s) to Slack.")
                for alert_msg in unique_alerts: send_slack_message(alert_msg); await asyncio.sleep(1)
            else: # No individual alerts from successful scrapes. Now check if all scrapes were successful.
                accounts_that_had_errors = len(df) - successful_scrape_mask.sum()
                if accounts_that_had_errors == 0: # All accounts were scraped successfully AND no individual alerts
                    logging.info("No individual data-driven alerts. All accounts scraped successfully. Sending 'All Clear' message.")
                    all_clear_message = f"🎉 Good news! Your friendly Usage Watchdog just completed its {now.strftime('%I:%M %p')} rounds, and all accounts are A-OK! :dog2::shield: Time to relax and enjoy the peace of mind! ✨"
                    send_slack_message(all_clear_message)
                else: # No individual alerts from successful scrapes, BUT some other accounts failed to scrape
                    logging.info(f"No specific data-driven alerts for successfully scraped accounts. {accounts_that_had_errors} account(s) had scraping issues.")
                    failed_accounts_details_list = []
                    for index, failed_row in df[~successful_scrape_mask].iterrows(): # Iterate over rows that ARE errors
                        failed_accounts_details_list.append(f"  - *{failed_row['Store']}* ({failed_row['Number']})")
                    failed_accounts_str = "\n".join(failed_accounts_details_list)
                    mixed_status_message = (
                        f":information_source: Usage check at {now.strftime('%I:%M %p')}:\n"
                        f"- No specific Low GB or Renewal/Low Balance alerts for successfully checked accounts.\n"
                        f"- However, data retrieval issues were encountered for the following *{accounts_that_had_errors} account(s)*:\n"
                        f"{failed_accounts_str}\n"
                        f"📧 Please check your email for the full report, which includes details for all accounts."
                    )
                    send_slack_message(mixed_status_message)
    
    # --- Prepare DataFrame for Excel Output ---
    df_for_excel = df.copy() 
    df_for_excel["Balance"] = df["Balance Numeric"]
    df_for_excel["Renewal Cost"] = df["Renewal Cost Numeric"]
    df_for_excel["Add-ons Price"] = df["Add-ons Price Numeric"]
    df_for_excel["Total Cost"] = df["Total Cost Numeric"]
    
    final_columns = ["Store", "Number", "Balance", "Main Quota", "Remaining", "Add-ons", "Add-ons Price", "Renewal Cost", "Total Cost", "Renewal Date"]
    df_for_excel = df_for_excel[final_columns]
    logging.info("DataFrame prepared for Excel export.")
    excel_path = "usage_report.xlsx"
    df_for_excel.to_excel(excel_path, index=False, sheet_name="Usage")
    logging.info(f"Data exported to {excel_path}.")

    # --- Styling with openpyxl ---
    wb = load_workbook(excel_path); ws = wb.active; center = Alignment(horizontal="center", vertical="center")
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells: cell.alignment = center
    header_row = ws[1]; col_letters = {}
    for cell in header_row:
        if cell.value: col_letters[cell.value] = cell.column_letter
    gb_format = '0" GB"'; egp_format = '#,##0.00" EGP"'
    if "Main Quota" in col_letters:
        for cell in ws[col_letters["Main Quota"]][1:]: cell.number_format = gb_format
    if "Remaining" in col_letters:
        for cell in ws[col_letters["Remaining"]][1:]: cell.number_format = gb_format
    currency_cols_for_formatting = ["Balance", "Add-ons Price", "Renewal Cost", "Total Cost"]
    for col_name in currency_cols_for_formatting:
        if col_name in col_letters:
            for cell in ws[col_letters[col_name]][1:]: cell.number_format = egp_format
        else: logging.warning(f"Column header '{col_name}' not found for EGP number formatting.")
            
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    if "Balance" in col_letters:
        balance_col_letter = col_letters["Balance"]
        for df_idx in range(len(df)): 
            excel_row_num = df_idx + 2
            try: 
                # Use ... Numeric columns from the MAIN df for this logic for styling
                balance_val = df.loc[df_idx, 'Balance Numeric'] 
                total_cost_val = df.loc[df_idx, 'Total Cost Numeric']
                
                # Check the original scraped 'Balance' string from the 'rows' list for this specific row
                # to avoid coloring cells that had critical scraping errors (which parse to 0)
                original_balance_str_for_style_check = rows[df_idx].get("Balance", "Error EGP") # Default to error if somehow missing

                if not ("Error EGP" in original_balance_str_for_style_check or "(Not Found)" in original_balance_str_for_style_check):
                    balance_cell = ws[f"{balance_col_letter}{excel_row_num}"]
                    epsilon = 1e-9
                    if balance_val < (total_cost_val - epsilon): balance_cell.fill = red_fill
                    elif abs(balance_val - total_cost_val) < epsilon: balance_cell.fill = yellow_fill
            except KeyError: pass 
            except Exception as e: logging.error(f"Error styling Balance cell {excel_row_num}: {e}")
        logging.info("Applied direct Balance cell styling.")

    if "Remaining" in col_letters:
        remaining_col_letter = col_letters["Remaining"]; last_row = ws.max_row
        dxf_red_gb = DifferentialStyle(fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        dxf_yellow_gb = DifferentialStyle(fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"))
        rule_red_gb = Rule(type="cellIs", operator="lessThan", formula=[str(LOW_REMAINING_RED_GB)], dxf=dxf_red_gb); rule_red_gb.stopIfTrue = True
        ws.conditional_formatting.add(f"{remaining_col_letter}2:{remaining_col_letter}{last_row}", rule_red_gb)
        rule_yellow_gb = Rule(type="cellIs", operator="lessThan", formula=[str(LOW_REMAINING_YELLOW_GB)], dxf=dxf_yellow_gb); rule_yellow_gb.stopIfTrue = True
        ws.conditional_formatting.add(f"{remaining_col_letter}2:{remaining_col_letter}{last_row}", rule_yellow_gb)
    logging.info("Excel file styled."); wb.save(excel_path)

    archive_daily_report()
    save_to_csv(df_for_excel)

    # --- Email Sending Logic ---
    now_for_dispatch = datetime.now() 
    target_daily_email_time = now_for_dispatch.replace(hour=12, minute=0, second=0, microsecond=0)
    interval_minutes_dispatch = 10 
    daily_email_window_start = target_daily_email_time - timedelta(minutes=interval_minutes_dispatch)
    daily_email_window_end = target_daily_email_time + timedelta(minutes=interval_minutes_dispatch)
    is_daily_email_time = daily_email_window_start <= now_for_dispatch <= daily_email_window_end

    # Use critically_failed_accounts_count determined earlier (counts "Error EGP" AND "Error" in Addons from rows)
    send_email_due_to_some_failures = critically_failed_accounts_count > 0 
    # (and critically_failed_accounts_count < len(rows) because the "all fail" case exits)
    email_subject = ""; email_contents = ""; should_send_this_email = False
    # now_for_dispatch = datetime.now().replace(hour=12, minute=5) # Test summary window

    if is_daily_email_time:
        should_send_this_email = True
        email_subject = f"📊 Daily Usage & Balance Report - {now_for_dispatch.strftime('%Y-%m-%d')}"
        email_contents = f"Please find today’s usage report attached (Generated around {now_for_dispatch.strftime('%I:%M %p')})."
        if send_email_due_to_some_failures: # It's daily report time AND there were some critical failures
             email_contents += f"\n\nNOTE: {critically_failed_accounts_count} account(s) encountered critical errors during data retrieval. Please check the report for details marked 'Error'."
        logging.info(f"Time {now_for_dispatch.strftime('%H:%M')} is in the daily email window.")
    
    elif send_email_due_to_some_failures: # Not daily report time, but there were critical failures
        should_send_this_email = True
        email_subject = f"⚠️ WE Usage Report - Data Retrieval Issues Detected - {now_for_dispatch.strftime('%Y-%m-%d %H:%M')}"
        email_contents = (
            f"{critically_failed_accounts_count} account(s) encountered critical errors during data retrieval. "
            f"The attached report contains details for all accounts, including error markers for these failed ones.\n\n"
            f"(Report generated on {now_for_dispatch.strftime('%Y-%m-%d at %I:%M %p')})."
        )
        logging.info(f"{critically_failed_accounts_count} account(s) failed critically. Preparing failure report email.")

    if should_send_this_email:
        if not TO_ADDRS: logging.warning("No email recipients configured. Skipping email.")
        elif not EMAIL_SENDER or not EMAIL_PASSWORD: logging.warning("Email sender or password not configured. Skipping email.")
        else:
            yag = None
            try:
                logging.info(f"Attempting to send email: \"{email_subject}\"")
                yag = yagmail.SMTP(EMAIL_SENDER, EMAIL_PASSWORD)
                yag.send(to=TO_ADDRS, subject=email_subject, contents=email_contents, attachments=[excel_path])
                logging.info(f"✅ Email sent to {TO_ADDRS} with {excel_path} attached.")
            except Exception as e: logging.error(f"Failed to send email: {e}", exc_info=True)
            finally:
                if yag:
                    try: yag.close()
                    except Exception as e: logging.error(f"Error closing yagmail connection: {e}")
    else:
        logging.info(f"No conditions met for sending email report at {now_for_dispatch.strftime('%H:%M')}.")

if __name__ == "__main__":
    asyncio.run(main())
   
